from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook

from app.factors.uk_2026_importer import (
    EXPECTED_HEADERS,
    FactorWorkbookValidationError,
    GreenhouseGasComponent,
    classify_greenhouse_gas_component,
    expected_headers,
    parse_uk_2023_flat_workbook,
    parse_uk_2024_flat_workbook,
    parse_uk_2025_flat_workbook,
    parse_uk_2026_flat_workbook,
)


def build_workbook(headers: tuple[str, ...] = EXPECTED_HEADERS) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Factors by Category"

    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=6, column=column, value=header)

    values = (
        "1_100_1000_15_1",
        "Scope 1",
        "Fuels",
        "Gaseous fuels",
        "Butane",
        None,
        None,
        "tonnes",
        "kg CO2e",
        "3033.38067",
    )
    for column, value in enumerate(values, start=1):
        worksheet.cell(row=7, column=column, value=value)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_parse_valid_workbook() -> None:
    parsed = parse_uk_2026_flat_workbook(build_workbook())

    assert parsed.total_data_rows == 1
    assert len(parsed.errors) == 0
    assert len(parsed.factors) == 1

    factor = parsed.factors[0]
    assert factor.source_factor_id == "1_100_1000_15_1"
    assert str(factor.factor_value) == "3033.38067"
    assert factor.greenhouse_gas_component == GreenhouseGasComponent.TOTAL_CO2E
    assert factor.factor_denominator_unit == "tonnes"
    assert factor.source_row_number == 7


def test_rejects_changed_header_schema() -> None:
    headers = list(EXPECTED_HEADERS)
    headers[-1] = "Unexpected Header"

    with pytest.raises(FactorWorkbookValidationError):
        parse_uk_2026_flat_workbook(build_workbook(tuple(headers)))


@pytest.mark.parametrize(
    ("label", "expected"),
    [
        ("kg CO2e", GreenhouseGasComponent.TOTAL_CO2E),
        ("kg CO2e of CO2 per unit", GreenhouseGasComponent.CO2),
        ("kg CO2e of CH4 per unit", GreenhouseGasComponent.CH4),
        ("kg CO2e of N2O per unit", GreenhouseGasComponent.N2O),
        ("custom factor", GreenhouseGasComponent.OTHER),
    ],
)
def test_classify_greenhouse_gas_component(
    label: str,
    expected: GreenhouseGasComponent,
) -> None:
    component, _ = classify_greenhouse_gas_component(label)

    assert component == expected


def test_skips_catalogue_row_with_no_published_factor() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Factors by Category"

    for column, header in enumerate(EXPECTED_HEADERS, start=1):
        worksheet.cell(row=6, column=column, value=header)

    values = (
        "1_101_1021_8_1",
        "Scope 1",
        "Fuels",
        "Liquid fuels",
        "Refinery miscellaneous",
        None,
        None,
        "litres",
        "kg CO2e",
        None,
    )
    for column, value in enumerate(values, start=1):
        worksheet.cell(row=7, column=column, value=value)

    output = BytesIO()
    workbook.save(output)
    workbook.close()

    parsed = parse_uk_2026_flat_workbook(output.getvalue())

    assert parsed.total_data_rows == 1
    assert parsed.skipped_unavailable_rows == 1
    assert parsed.factors == ()
    assert parsed.errors == ()


def test_parse_uk_2024_hvo_factors() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Factors by Category"

    for column, header in enumerate(expected_headers(2024), start=1):
        worksheet.cell(row=6, column=column, value=header)

    rows = (
        (
            "2_103_1036_8_1",
            "Scope 1",
            "Bioenergy",
            "Biofuel",
            "Biodiesel HVO",
            None,
            None,
            "litres",
            "kg CO2e",
            "0.03558",
        ),
        (
            "12_900_1036_8_1",
            "Scope 3",
            "WTT- bioenergy",
            "WTT- biofuel",
            "Biodiesel HVO",
            None,
            None,
            "litres",
            "kg CO2e",
            "0.559",
        ),
        (
            "99_103_1036_8_2",
            "Outside of Scopes",
            "Bioenergy",
            "Biofuel",
            "Biodiesel HVO",
            None,
            None,
            "litres",
            "kg CO2e of CO2 per unit",
            "2.43",
        ),
    )
    for row_number, values in enumerate(rows, start=7):
        for column, value in enumerate(values, start=1):
            worksheet.cell(row=row_number, column=column, value=value)

    output = BytesIO()
    workbook.save(output)
    workbook.close()

    parsed = parse_uk_2024_flat_workbook(output.getvalue())
    factors = {item.source_factor_id: item for item in parsed.factors}

    assert parsed.errors == ()
    assert str(factors["2_103_1036_8_1"].factor_value) == "0.03558"
    assert factors["2_103_1036_8_1"].lifecycle_boundary == "direct"
    assert str(factors["12_900_1036_8_1"].factor_value) == "0.559"
    assert factors["12_900_1036_8_1"].lifecycle_boundary == "well_to_tank"
    assert str(factors["99_103_1036_8_2"].factor_value) == "2.43"
    assert factors["99_103_1036_8_2"].greenhouse_gas_component == GreenhouseGasComponent.CO2


def test_parse_uk_2023_hvo_factors() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Factors by Category"

    for column, header in enumerate(expected_headers(2023), start=1):
        worksheet.cell(row=6, column=column, value=header)

    rows = (
        (
            "2_103_1036_8_1",
            "Scope 1",
            "Bioenergy",
            "Biofuel",
            "Biodiesel HVO",
            None,
            None,
            "litres",
            "kg CO2e",
            "0.03558",
        ),
        (
            "12_900_1036_8_1",
            "Scope 3",
            "WTT- bioenergy",
            "WTT- biofuel",
            "Biodiesel HVO",
            None,
            None,
            "litres",
            "kg CO2e",
            "0.27844",
        ),
        (
            "99_103_1036_8_2",
            "Outside of Scopes",
            "Outside of scopes",
            "Biofuel",
            "Biodiesel HVO",
            None,
            None,
            "litres",
            "kg CO2e of CO2 per unit",
            "2.43",
        ),
    )
    for row_number, values in enumerate(rows, start=7):
        for column, value in enumerate(values, start=1):
            worksheet.cell(row=row_number, column=column, value=value)

    output = BytesIO()
    workbook.save(output)
    workbook.close()

    parsed = parse_uk_2023_flat_workbook(output.getvalue())
    factors = {item.source_factor_id: item for item in parsed.factors}

    assert parsed.errors == ()
    assert str(factors["2_103_1036_8_1"].factor_value) == "0.03558"
    assert str(factors["12_900_1036_8_1"].factor_value) == "0.27844"
    assert str(factors["99_103_1036_8_2"].factor_value) == "2.43"
    assert factors["99_103_1036_8_2"].greenhouse_gas_component == GreenhouseGasComponent.CO2


def test_uk_2023_parser_rejects_2024_headers() -> None:
    with pytest.raises(FactorWorkbookValidationError):
        parse_uk_2023_flat_workbook(build_workbook(expected_headers(2024)))


def test_uk_2024_parser_rejects_2026_headers() -> None:
    with pytest.raises(FactorWorkbookValidationError):
        parse_uk_2024_flat_workbook(build_workbook())


def test_parse_uk_2025_flat_workbook() -> None:
    parsed = parse_uk_2025_flat_workbook(build_workbook(expected_headers(2025)))

    assert parsed.errors == ()
    assert parsed.total_data_rows == 1
    assert str(parsed.factors[0].factor_value) == "3033.38067"
