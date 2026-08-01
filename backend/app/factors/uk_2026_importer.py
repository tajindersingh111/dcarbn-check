from __future__ import annotations

import hashlib
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, BinaryIO

from openpyxl import load_workbook

from app.models.emission_factor import GreenhouseGasComponent


WORKSHEET_NAME = "Factors by Category"
HEADER_ROW = 6
EXPECTED_HEADERS = (
    "ID",
    "Scope",
    "Level 1",
    "Level 2",
    "Level 3",
    "Level 4",
    "Column Text",
    "UOM",
    "GHG/Unit",
    "GHG Conversion Factor 2026",
)


class FactorWorkbookValidationError(ValueError):
    pass


@dataclass(frozen=True, slots=True)
class ParsedFactorRow:
    source_factor_id: str
    scope: str
    level_1: str
    level_2: str | None
    level_3: str | None
    level_4: str | None
    column_text: str | None
    activity_unit: str
    factor_unit_text: str
    greenhouse_gas_component: GreenhouseGasComponent
    greenhouse_gas_label: str
    factor_value: Decimal
    factor_numerator_unit: str
    factor_denominator_unit: str
    lifecycle_boundary: str | None
    source_row_number: int
    raw_source_data: dict[str, Any]


@dataclass(frozen=True, slots=True)
class RowParseError:
    row_number: int
    error_code: str
    message: str
    raw_source_data: dict[str, Any]


@dataclass(frozen=True, slots=True)
class ParsedWorkbook:
    source_sha256: str
    factors: tuple[ParsedFactorRow, ...]
    errors: tuple[RowParseError, ...]
    total_data_rows: int
    skipped_unavailable_rows: int


def calculate_sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _optional_text(value: object) -> str | None:
    cleaned = _clean_text(value)
    return cleaned or None


def classify_greenhouse_gas_component(
    factor_unit_text: str,
) -> tuple[GreenhouseGasComponent, str]:
    normalized = " ".join(factor_unit_text.lower().split())

    if "of co2 per unit" in normalized:
        return GreenhouseGasComponent.CO2, "CO2"
    if "of ch4 per unit" in normalized:
        return GreenhouseGasComponent.CH4, "CH4"
    if "of n2o per unit" in normalized:
        return GreenhouseGasComponent.N2O, "N2O"
    if normalized in {"kg co2e", "kgco2e"}:
        return GreenhouseGasComponent.TOTAL_CO2E, "CO2e"
    return GreenhouseGasComponent.OTHER, factor_unit_text


def infer_lifecycle_boundary(scope: str, level_1: str, column_text: str | None) -> str | None:
    joined = " ".join(
        part for part in (scope, level_1, column_text or "") if part
    ).lower()

    if "well-to-tank" in joined or "wtt" in joined:
        return "well_to_tank"
    if "transmission and distribution" in joined or "t&d" in joined:
        return "transmission_and_distribution"
    if "scope 1" in joined:
        return "direct"
    if "scope 2" in joined:
        return "purchased_energy"
    if "scope 3" in joined:
        return "indirect_value_chain"
    return None


def parse_uk_2026_flat_workbook(content: bytes) -> ParsedWorkbook:
    workbook = load_workbook(
        filename=BytesIO(content),
        read_only=True,
        data_only=True,
    )

    try:
        if WORKSHEET_NAME not in workbook.sheetnames:
            raise FactorWorkbookValidationError(
                f"Required worksheet '{WORKSHEET_NAME}' was not found."
            )

        worksheet = workbook[WORKSHEET_NAME]
        actual_headers = tuple(
            _clean_text(worksheet.cell(row=HEADER_ROW, column=index).value)
            for index in range(1, len(EXPECTED_HEADERS) + 1)
        )
        if actual_headers != EXPECTED_HEADERS:
            raise FactorWorkbookValidationError(
                "The UK 2026 flat-format header row does not match the expected "
                f"schema. Expected {EXPECTED_HEADERS!r}; received {actual_headers!r}."
            )

        factors: list[ParsedFactorRow] = []
        errors: list[RowParseError] = []
        total_data_rows = 0
        skipped_unavailable_rows = 0

        for row_number, values in enumerate(
            worksheet.iter_rows(
                min_row=HEADER_ROW + 1,
                max_col=len(EXPECTED_HEADERS),
                values_only=True,
            ),
            start=HEADER_ROW + 1,
        ):
            if not any(value is not None and _clean_text(value) for value in values):
                continue

            total_data_rows += 1
            raw = {
                header: value
                for header, value in zip(EXPECTED_HEADERS, values, strict=True)
            }

            if _clean_text(raw["Scope"]).upper() == "END":
                continue

            try:
                source_factor_id = _clean_text(raw["ID"])
                scope = _clean_text(raw["Scope"])
                level_1 = _clean_text(raw["Level 1"])
                activity_unit = _clean_text(raw["UOM"])
                factor_unit_text = _clean_text(raw["GHG/Unit"])
                factor_value_text = _clean_text(raw["GHG Conversion Factor 2026"])

                missing_descriptors = [
                    field
                    for field, value in (
                        ("ID", source_factor_id),
                        ("Scope", scope),
                        ("Level 1", level_1),
                        ("UOM", activity_unit),
                        ("GHG/Unit", factor_unit_text),
                    )
                    if not value
                ]
                if missing_descriptors:
                    raise FactorWorkbookValidationError(
                        "Required descriptor values are missing: "
                        f"{', '.join(missing_descriptors)}."
                    )

                if not factor_value_text:
                    skipped_unavailable_rows += 1
                    continue

                try:
                    factor_value = Decimal(factor_value_text)
                except InvalidOperation as exc:
                    raise FactorWorkbookValidationError(
                        "GHG Conversion Factor 2026 must be a valid decimal."
                    ) from exc

                component, greenhouse_gas_label = classify_greenhouse_gas_component(
                    factor_unit_text
                )
                column_text = _optional_text(raw["Column Text"])

                factors.append(
                    ParsedFactorRow(
                        source_factor_id=source_factor_id,
                        scope=scope,
                        level_1=level_1,
                        level_2=_optional_text(raw["Level 2"]),
                        level_3=_optional_text(raw["Level 3"]),
                        level_4=_optional_text(raw["Level 4"]),
                        column_text=column_text,
                        activity_unit=activity_unit,
                        factor_unit_text=factor_unit_text,
                        greenhouse_gas_component=component,
                        greenhouse_gas_label=greenhouse_gas_label,
                        factor_value=factor_value,
                        factor_numerator_unit="kg CO2e",
                        factor_denominator_unit=activity_unit,
                        lifecycle_boundary=infer_lifecycle_boundary(
                            scope,
                            level_1,
                            column_text,
                        ),
                        source_row_number=row_number,
                        raw_source_data=raw,
                    )
                )
            except FactorWorkbookValidationError as exc:
                errors.append(
                    RowParseError(
                        row_number=row_number,
                        error_code="invalid_factor_row",
                        message=str(exc),
                        raw_source_data=raw,
                    )
                )

        if not factors and not errors:
            raise FactorWorkbookValidationError(
                "No factor rows were found in the workbook."
            )

        return ParsedWorkbook(
            source_sha256=calculate_sha256(content),
            factors=tuple(factors),
            errors=tuple(errors),
            total_data_rows=total_data_rows,
            skipped_unavailable_rows=skipped_unavailable_rows,
        )
    finally:
        workbook.close()


def read_binary_stream(stream: BinaryIO, maximum_bytes: int = 25_000_000) -> bytes:
    content = stream.read(maximum_bytes + 1)
    if len(content) > maximum_bytes:
        raise FactorWorkbookValidationError(
            f"Workbook exceeds the {maximum_bytes} byte upload limit."
        )
    if not content:
        raise FactorWorkbookValidationError("Workbook is empty.")
    return content
