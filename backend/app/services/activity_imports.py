from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile, ZipFile

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

MAX_WORKBOOK_BYTES = 5_000_000
MAX_WORKBOOK_UNCOMPRESSED_BYTES = 20_000_000
MAX_WORKBOOK_ARCHIVE_ENTRIES = 100
MAX_ACTIVITY_ROWS = 500
MAX_ACTIVITY_COLUMNS = 50


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, float):
        return format(value, ".15g")
    return str(value).strip()


def _validate_workbook_archive(content: bytes) -> None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_WORKBOOK_ARCHIVE_ENTRIES:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="The workbook contains too many internal files.",
                )
            if sum(entry.file_size for entry in entries) > MAX_WORKBOOK_UNCOMPRESSED_BYTES:
                raise HTTPException(
                    status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                    detail="The expanded workbook exceeds the 20 MB safety limit.",
                )
            if any(
                entry.filename.startswith(("/", "\\"))
                or ".." in Path(entry.filename).parts
                for entry in entries
            ):
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="The workbook contains an unsafe internal path.",
                )
    except BadZipFile as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="The workbook is not a valid .xlsx file.",
        ) from exc


async def parse_activity_workbook(
    upload: UploadFile,
) -> tuple[list[str], list[list[str]]]:
    filename = upload.filename or ""
    if Path(filename).suffix.lower() != ".xlsx":
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Choose an Excel .xlsx workbook.",
        )

    content = await upload.read(MAX_WORKBOOK_BYTES + 1)
    if len(content) > MAX_WORKBOOK_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="The workbook exceeds the 5 MB upload limit.",
        )
    if not content:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="The workbook is empty.",
        )

    _validate_workbook_archive(content)

    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except (InvalidFileException, OSError, ValueError) as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="The workbook is not a valid .xlsx file.",
        ) from exc

    populated_sheets: list[tuple[list[str], list[list[str]]]] = []
    try:
        for worksheet in workbook.worksheets:
            if worksheet.max_column > MAX_ACTIVITY_COLUMNS:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail=(
                        f"Worksheet “{worksheet.title}” exceeds the "
                        f"{MAX_ACTIVITY_COLUMNS}-column limit."
                    ),
                )
            if worksheet.max_row > MAX_ACTIVITY_ROWS + 1:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail=(
                        f"Worksheet “{worksheet.title}” exceeds the "
                        f"{MAX_ACTIVITY_ROWS}-row import limit."
                    ),
                )

            parsed_rows: list[list[str]] = []
            for row in worksheet.iter_rows():
                if any(cell.data_type == "f" for cell in row):
                    raise HTTPException(
                        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                        detail=(
                            f"Worksheet “{worksheet.title}” contains a formula. "
                            "Replace formulas with their displayed values before upload."
                        ),
                    )
                values = [_cell_text(cell.value) for cell in row]
                while values and not values[-1]:
                    values.pop()
                if any(values):
                    parsed_rows.append(values)

            if parsed_rows:
                headers, *rows = parsed_rows
                populated_sheets.append((headers, rows))
    finally:
        workbook.close()

    if not populated_sheets:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="The workbook does not contain an activity table.",
        )
    if len(populated_sheets) > 1:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Use one populated worksheet per upload.",
        )

    headers, rows = populated_sheets[0]
    if not headers:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="The first row must contain column headings.",
        )
    return headers, rows
